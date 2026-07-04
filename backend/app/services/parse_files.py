from __future__ import annotations

import csv
import io
import re
from typing import Any

from openpyxl import load_workbook

from app.models import FieldMapping, SurveyPoint


RawTable = tuple[list[str], list[dict[str, Any]]]


FIELD_ALIASES = {
    "pointId": ["point_id", "pointid", "point", "id", "no", "pn", "点号", "点名", "编号"],
    "east": ["east", "e", "x", "easting", "横坐标", "东坐标", "东", "x坐标"],
    "north": ["north", "n", "y", "northing", "纵坐标", "北坐标", "北", "y坐标"],
    "height": ["height", "h", "z", "elevation", "高程", "标高", "高"],
    "note": ["note", "remark", "remarks", "description", "备注", "说明"],
    "code": ["code", "cass", "feature_code", "地物编码", "编码", "代码"],
}


def parse_uploaded_file(filename: str, content: bytes) -> RawTable:
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if suffix == "csv":
        return parse_csv(content)
    if suffix == "xlsx":
        return parse_xlsx(content)
    if suffix == "dat":
        return parse_dat(content)
    raise ValueError("Only CSV, XLSX, and DAT files are supported.")


def parse_csv(content: bytes) -> RawTable:
    text = decode_text(content)
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        raise ValueError("CSV header is required.")
    raw_columns = [str(column) for column in reader.fieldnames]
    columns = [column.strip() for column in raw_columns]
    rows = [normalize_csv_row(row, raw_columns, columns) for row in reader if any(string_value(v) for v in row.values())]
    return columns, rows


def parse_xlsx(content: bytes) -> RawTable:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("XLSX has no data.")

    header = [cell_to_string(value) for value in rows[0]]
    if not any(header):
        raise ValueError("XLSX first row must contain headers.")
    columns = [value or f"col_{index + 1}" for index, value in enumerate(header)]

    data_rows: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(cell_to_string(value) for value in row):
            continue
        item = {}
        for index, column in enumerate(columns):
            item[column] = cell_to_string(row[index]) if index < len(row) else ""
        data_rows.append(item)
    return columns, data_rows


def parse_dat(content: bytes) -> RawTable:
    text = decode_text(content)
    columns = ["point_id", "empty", "east", "north", "height", "note", "code"]
    rows: list[dict[str, Any]] = []

    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue

        parts = split_dat_line(line)
        if len(parts) >= 5 and parts[1] == "":
            point_id, _, east, north, height, *rest = parts
            code = rest[1] if len(rest) > 1 else ""
        elif len(parts) == 5 and is_number_text(parts[1]):
            point_id, code, east, north, height = parts
            rest = []
        elif len(parts) >= 4:
            point_id, east, north, height, *rest = parts
            code = rest[1] if len(rest) > 1 else ""
        else:
            raise ValueError(f"Invalid DAT row: {raw_line}")

        rows.append(
            {
                "point_id": point_id,
                "empty": "",
                "east": east,
                "north": north,
                "height": height,
                "note": rest[0] if rest else "",
                "code": code,
            }
        )
    return columns, rows


def split_dat_line(line: str) -> list[str]:
    if "," in line or "，" in line:
        return [part.strip() for part in re.split(r"[,，]", line)]
    return [part.strip() for part in re.split(r"\s+", line) if part.strip()]


def detect_mapping(columns: list[str]) -> dict[str, str | None]:
    detected: dict[str, str | None] = {}
    normalized = {normalize_name(column): column for column in columns}
    for field, aliases in FIELD_ALIASES.items():
        detected[field] = None
        for alias in aliases:
            key = normalize_name(alias)
            if key in normalized:
                detected[field] = normalized[key]
                break
    return detected


def rows_to_points(rows: list[dict[str, Any]], mapping: FieldMapping) -> list[SurveyPoint]:
    points: list[SurveyPoint] = []
    for index, row in enumerate(rows, start=1):
        point_id = string_value(row.get(mapping.pointId))
        if not point_id:
            raise ValueError(f"Row {index} is missing point id.")

        east = parse_float(row.get(mapping.east), f"Row {index} east")
        north = parse_float(row.get(mapping.north), f"Row {index} north")
        height = parse_optional_float(row.get(mapping.height), f"Row {index} height") if mapping.height else None
        note = string_value(row.get(mapping.note)) if mapping.note else ""
        code = string_value(row.get(mapping.code)) if mapping.code else ""
        points.append(SurveyPoint(id=point_id, east=east, north=north, height=height, note=note, code=code))
    return points


def decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def normalize_row(row: dict[str, Any], columns: list[str]) -> dict[str, Any]:
    return {column: cell_to_string(row.get(column)) for column in columns}


def normalize_csv_row(row: dict[str, Any], raw_columns: list[str], columns: list[str]) -> dict[str, Any]:
    return {column: cell_to_string(row.get(raw_column)) for raw_column, column in zip(raw_columns, columns)}


def normalize_name(value: str) -> str:
    return re.sub(r"[\s_\-（）()]+", "", str(value).strip().lower())


def cell_to_string(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def string_value(value: Any) -> str:
    return "" if value is None else str(value).strip()


def is_number_text(value: Any) -> bool:
    text = string_value(value)
    if not text:
        return False
    try:
        float(text)
        return True
    except ValueError:
        return False


def parse_float(value: Any, label: str) -> float:
    text = string_value(value)
    if not text:
        raise ValueError(f"{label} is missing.")
    try:
        return float(text)
    except ValueError as exc:
        raise ValueError(f"{label} is not a number: {text}") from exc


def parse_optional_float(value: Any, label: str) -> float | None:
    text = string_value(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError as exc:
        raise ValueError(f"{label} is not a number: {text}") from exc
